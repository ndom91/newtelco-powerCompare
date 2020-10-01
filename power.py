#!/usr/bin/env python3

import sys
import subprocess
import getopt
import datetime
import openpyxl
import MySQLdb
import psycopg2
import pandas as pd
import numpy as np
import argparse
import array
import math
import dbconfig as cfg
import datetime
import arrow
# import pynetbox
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Font, Side, Alignment, Fill, PatternFill, NamedStyle
from openpyxl.worksheet import worksheet
from openpyxl import Workbook

# Helper Functions


def truncate(number, digits) -> float:
    stepper = pow(10.0, digits)
    return math.trunc(stepper * number) / stepper


def move_cell(source_cell, coord, tgt):
    tgt[coord].value = source_cell.value
    if source_cell.has_style:
        tgt[coord]._style = copy(source_cell._style)

    del source_cell.parent._cells[(source_cell.row, source_cell.col_idx)]

    return tgt[coord]


def get_column_letter(n):
    convertString = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    base = 26
    i = n - 1

    if i < base:
        return convertString[i]
    else:
        return get_column_letter(i//base) + convertString[i % base]


def getPsql(date):
    # PSQL Connection (localhost)
    # Query Netbox Database for Contract / Rack allowed power usage
    try:
        # nb = pynetbox.api(
        #     'https://racks.newtelco.de',
        #     token='796945c353b2508588b52f783c81110c381b34ad'
        # )
        # rackInfo = list()
        # for i in nb.dcim.racks.filter(name__icontains="R"):
        #   rackInfo.append(
        #     dict(
        #       name=i.name,
        #       status=i.status
        #     )
        #   )
        #   print(i)
        #   print(i.fields())
        #   # print(i.status)
        #   # rackInfo.append()

        # # print(rackInfo)

        psql_db = cfg.psql['database']
        psql_user = cfg.psql['user']
        psql_host = cfg.psql['host']
        psql_pw = cfg.psql['password']
        psql_port = cfg.psql['port']

        conn = psycopg2.connect(
            dbname=psql_db, user=psql_user, host=psql_host, password=psql_pw, port=psql_port)

        cursor = conn.cursor()

        cursor.execute(
            """SELECT dcim_rack.name FROM dcim_rack LEFT JOIN extras_customfieldvalue ON dcim_rack.id = extras_customfieldvalue.obj_id WHERE dcim_rack.site_id = '1';""")
        psqlRows = cursor.fetchall()

        # 9 - Contract Power DC
        cursor.execute("""SELECT dcim_rack.name, extras_customfieldvalue.serialized_value FROM dcim_rack LEFT JOIN extras_customfieldvalue ON dcim_rack.id = extras_customfieldvalue.obj_id WHERE dcim_rack.site_id = '1' AND extras_customfieldvalue.field_id = '9';""")
        fieldidDC = cursor.fetchall()

        # 9 - Contract Power AC
        cursor.execute("""SELECT dcim_rack.name, extras_customfieldvalue.serialized_value FROM dcim_rack LEFT JOIN extras_customfieldvalue ON dcim_rack.id = extras_customfieldvalue.obj_id WHERE dcim_rack.site_id = '1' AND extras_customfieldvalue.field_id = '2';""")
        fieldidAC = cursor.fetchall()

        # 10 - Contract Number
        cursor.execute("""SELECT dcim_rack.name, extras_customfieldvalue.serialized_value as contract FROM dcim_rack LEFT JOIN extras_customfieldvalue ON dcim_rack.id = extras_customfieldvalue.obj_id WHERE dcim_rack.site_id = '1' AND extras_customfieldvalue.field_id = '10';""")
        fieldidContract = cursor.fetchall()

        # 11 - Counter A Number
        cursor.execute("""SELECT dcim_rack.name, extras_customfieldvalue.serialized_value as counterA FROM dcim_rack LEFT JOIN extras_customfieldvalue ON dcim_rack.id = extras_customfieldvalue.obj_id WHERE dcim_rack.site_id = '1' AND extras_customfieldvalue.field_id = '11' AND extras_customfieldvalue.serialized_value != '';""")
        fieldidCounterA = cursor.fetchall()

        # 11 - Counter B Number
        cursor.execute("""SELECT dcim_rack.name, extras_customfieldvalue.serialized_value as counterB FROM dcim_rack LEFT JOIN extras_customfieldvalue ON dcim_rack.id = extras_customfieldvalue.obj_id WHERE dcim_rack.site_id = '1' AND extras_customfieldvalue.field_id = '8' AND extras_customfieldvalue.serialized_value != '';""")
        fieldidCounterB = cursor.fetchall()

        cursor.close()
        conn.close()

        return [psqlRows, fieldidAC, fieldidDC, fieldidContract, fieldidCounterA, fieldidCounterB]

    except Exception as e:
        print("Uh oh, can't connect. Invalid dbname, user or password?")
        print(e)


def getMysql(date):
    # MySQL Connection
    # Query CRM MySQL DB for current rack usage
    try:
        connection = MySQLdb.connect(host=cfg.mysql['host'],
                                     port=cfg.mysql['port'],
                                     user=cfg.mysql['user'],
                                     passwd=cfg.mysql['passwd'],
                                     db=cfg.mysql['db'])

        counterValues = connection.cursor()

        reqMonth = date[-2:]
        reqYear = date[:-2]

        nowm0 = datetime.datetime(int(reqYear), int(reqMonth), 2)
        nowm1 = nowm0 - datetime.timedelta(days=30)
        nowm2 = nowm1 - datetime.timedelta(days=30)

        yearMonth0 = nowm0.strftime("%Y%m")
        yearMonth1 = nowm1.strftime("%Y%m")
        yearMonth2 = nowm2.strftime("%Y%m")

        q = "select powerCounters.serialNo, CONCAT(powerCounters.rNumber, ' ',company.company) as name, powerCounters.rNumber, powerCounterValues.sortDateTime, powerCounterValues.diff FROM powerCounters left join powerCounterValues on powerCounters.id = powerCounterValues.counterId left join company on powerCounters.companyId = company.id WHERE powerCounterValues.sortDateTime = %(date)s;"
        params = {'date': yearMonth0}
        counterValues.execute(q, params)
        mysqlRows = counterValues.fetchall()

        q = "select powerCounters.serialNo, powerCounterValues.sortDateTime, powerCounterValues.diff FROM powerCounters left join powerCounterValues on powerCounters.id = powerCounterValues.counterId left join company on powerCounters.companyId = company.id WHERE powerCounterValues.sortDateTime = %(date)s;"
        params = {'date': yearMonth1}
        counterValues.execute(q, params)
        mysqlRowsm1 = counterValues.fetchall()

        q = "select powerCounters.serialNo, powerCounterValues.sortDateTime, powerCounterValues.diff FROM powerCounters left join powerCounterValues on powerCounters.id = powerCounterValues.counterId left join company on powerCounters.companyId = company.id WHERE powerCounterValues.sortDateTime = %(date)s;"
        params = {'date': yearMonth2}
        counterValues.execute(q, params)
        mysqlRowsm2 = counterValues.fetchall()

        counterValues.close()
        connection.close()

        return [mysqlRows, mysqlRowsm1, mysqlRowsm2]

    except Exception as e:
        print("Uh oh, can't connect. Invalid dbname, user or password?")
        print(e)


def compare(date):

    psqlRows, fieldidAC, fieldidDC, fieldidContract, fieldidCounterA, fieldidCounterB = getPsql(date)
    mysqlRows, mysqlRowsm1, mysqlRowsm2 = getMysql(date)

    mysqlArr = np.asarray(mysqlRows)

    # np.set_printoptions(threshold=sys.maxsize)
    # print('mysqlArr')
    # print(mysqlArr[:5])
    # print(mysqlArr)
    # [['30111782' 'R2701 Remba Telecom Ltd' 'R2701' 201911 320]
    # ['30103897' 'R2702 NewTelco South Africa' 'R2702' 201911 1212]
    # ['6697500' 'R2703 NewTelco South Africa' 'R2703' 201911 1116]
    # ['6295839' 'R2705 Cinia Oy' 'R2705' 201911 0]
    # ['30111783' 'R2706 Avalon Telecom SIA' 'R2706' 201911 324]]

    mysqlm1Arr = np.asarray(mysqlRowsm1)
    # print('mysqlm1Arr')
    # print(mysqlm1Arr[:5])
    # [['30111782' '201910' '330']
    # ['30103897' '201910' '1251']
    # ['6697500' '201910' '1159']
    # ['6295839' '201910' '0']
    # ['30111783' '201910' '348']]

    mysqlm2Arr = np.asarray(mysqlRowsm2)
    # print('mysqlm2Arr')
    # print(mysqlm2Arr[:5])
    # [['30111782' '201909' '318']
    # ['30103897' '201909' '1209']
    # ['6697500' '201909' '1123']
    # ['6295839' '201909' '0']
    # ['30111783' '201909' '336']]

    psqlArr = np.asarray(psqlRows)
    # print('psqlArr')
    # print(psqlArr[:5])
    # [['R1947 Rascom']
    # ['R1915 TATA']
    # ['R1916 TATA']
    # ['R1916 TATA']
    # ['R1916 TATA']]

    fieldidACArr = np.asarray(fieldidAC)
    # print('fieldidACArr')
    # print(fieldidACArr[:5])
    # [['R1915 TATA' '7000']
    # ['R1916 TATA' '7000']
    # ['R1920 TATA' '7000']
    # ['R1917 TATA' '3500']
    # ['R1918 TATA' '10500']]

    fieldidDCArr = np.asarray(fieldidDC)
    # print('fieldidDCArr')
    # print(fieldidDCArr[:5])
    # [['R1955 Megafon' '70']
    # ['R1956 Megafon' '60']
    # ['R1952 Vodafone' '60']
    # ['R2204 Oblcom' '6']
    # ['Vodafone B02.104' '703']]

    fieldidContractArr = np.asarray(fieldidContract)
    # print('fieldidContractArr')
    # print(fieldidContractArr[:5])
    # [['R1947 Rascom' '140475']
    # ['R1916 TATA' '140372']
    # ['R1921 TATA' '140372']
    # ['R1917 TATA' '140372']
    # ['R1918 TATA' '140372']]

    fieldidCounterAArr = np.asarray(fieldidCounterA)
    # print('fieldidCounterAArr')
    # print(fieldidCounterAArr[:5])
    # [['R1916 TATA' '30101763']
    # ['R1920 TATA' '12100019-2']
    # ['R1917 TATA' '30103558']
    # ['R1926 Press TV' '12100000-1']
    # ['R1930 Kavir' '12100009-1']]

    fieldidCounterBArr = np.asarray(fieldidCounterB)
    # print('fieldidCounterBArr')
    # print(fieldidCounterBArr[:5])
    # [['R1916 TATA' '30101722']
    # ['R1920 TATA' '06100000-2']
    # ['R1917 TATA' '12100021-2']
    # ['R1930 Kavir' '11100002-1']
    # ['R1939 Silknet' '11100004-2']]

    mysqlDF = pd.DataFrame({'Counter': mysqlArr[:, 0], 'name': mysqlArr[:, 1],
                            'Rack': mysqlArr[:, 2], 'Month': mysqlArr[:, 3], 'Usage': mysqlArr[:, 4]},)
    mysqlDF['Usage'] = mysqlDF['Usage'].infer_objects()
    # print('mysqlDF')
    # pd.set_option('display.max_rows', None)
    # print(mysqlDF)
    #     Counter                        name   Rack   Month  Usage
    # 0  30100767  R2707 MTBC Telecom Limited  R2707  201911  715.0
    # 1  30100731             R2708 IP-MAX SA  R2708  201911  387.0
    # 2  30100650         R2708 NewTelco GmbH  R2708  201911   73.0
    # 3  30100747            R2715 Fortex ZAO  R2715  201911  178.0
    # 4  30100761               R2719 Silknet  R2719  201911  209.0

    mysqlm1DF = pd.DataFrame(
        {'Counter': mysqlm1Arr[:, 0], 'Month-1': mysqlm1Arr[:, 1], 'Usage M-1': mysqlm1Arr[:, 2]})
    mysqlm1DF['Usage M-1'] = mysqlm1DF['Usage M-1'].infer_objects()
    # print('mysqlm1DF')
    # print(mysqlm1DF[:5])
    #     Counter Month-1 Usage M-1
    # 0  30111782  201910       330
    # 1  30103897  201910      1251
    # 2   6697500  201910      1159
    # 3   6295839  201910         0
    # 4  30111783  201910       348

    mysqlm2DF = pd.DataFrame(
        {'Counter': mysqlm2Arr[:, 0], 'Month-2': mysqlm2Arr[:, 1], 'Usage M-2': mysqlm2Arr[:, 2]})
    mysqlm2DF['Usage M-2'] = mysqlm2DF['Usage M-2'].infer_objects()
    # print('mysqlm2DF')
    # print(mysqlm2DF[:5])
    #     Counter Month-2 Usage M-2
    # 0  30111782  201909       318
    # 1  30103897  201909      1209
    # 2   6697500  201909      1123
    # 3   6295839  201909         0
    # 4  30111783  201909       336

    # for index, row in mysqlDF.iterrows():
    #     row['name'] = row['name'].replace(r' A-Feed', '')
    #     row['name'] = row['name'].replace(r' B-Feed', '')

    fidACDF = pd.DataFrame(
        {'name': fieldidACArr[:, 0], 'AC': fieldidACArr[:, 1]})
    # print('fidACDF')
    # print(fidACDF.head())
    #          name     AC
    # 0  R1915 TATA   7000
    # 1  R1916 TATA   7000
    # 2  R1920 TATA   7000
    # 3  R1917 TATA   3500
    # 4  R1918 TATA  10500

    fidDCDF = pd.DataFrame(
        {'name': fieldidDCArr[:, 0], 'DC': fieldidDCArr[:, 1]})
    # print('fidDCDF')
    # print(fidDCDF.head())
    #                name   DC
    # 0     R1955 Megafon   70
    # 1     R1956 Megafon   60
    # 2    R1952 Vodafone   60
    # 3      R2204 Oblcom    6
    # 4  Vodafone B02.104  703

    fidContractDF = pd.DataFrame(
        {'name': fieldidContractArr[:, 0], 'Contract': fieldidContractArr[:, 1]})
    # print('fidContractDF')
    # print(fidContractDF.head())
    #        name      Contract
    # 0  R1947 Rascom   140475
    # 1    R1916 TATA   140372
    # 2    R1921 TATA   140372
    # 3    R1917 TATA   140372
    # 4    R1918 TATA   140372

    fidCounterADF = pd.DataFrame(
        {'name': fieldidCounterAArr[:, 0], 'CounterB': fieldidCounterAArr[:, 1]})
    # print('fidCounterADF')
    # print(fidCounterADF.head())
    #         name       CounterB
    # 0      R1916 TATA    30101763
    # 1      R1920 TATA  12100019-2
    # 2      R1917 TATA    30103558
    # 3  R1926 Press TV  12100000-1
    # 4     R1930 Kavir  12100009-1

    fidCounterBDF = pd.DataFrame(
        {'name': fieldidCounterBArr[:, 0], 'CounterA': fieldidCounterBArr[:, 1]})
    # print('fidCounterBDF')
    # print(fidCounterBDF.head())
    #         name      CounterA
    # 0     R1916 TATA    30101722
    # 1     R1920 TATA  06100000-2
    # 2     R1917 TATA  12100021-2
    # 3    R1930 Kavir  11100002-1
    # 4  R1939 Silknet  11100004-2

    #####################################################
    # At this point we have all psql rows where counterA and counterB are not empty..
    #####################################################

    # psqlDF = pd.DataFrame({'name':psqlArr[:,0]})

    # merge1 = pd.merge(mysqlDF, psqlDF, left_on='name', right_on='name', how='left')

    merge4 = pd.merge(fidDCDF, fidACDF, left_on='name',
                      right_on='name', how='right')
    merge5 = pd.merge(merge4, fidContractDF, left_on='name',
                      right_on='name', how='right')
    merge6 = pd.merge(merge5, fidCounterADF, left_on='name',
                      right_on='name', how='right')
    merge6A = pd.merge(merge6, fidCounterBDF, left_on='name',
                       right_on='name', how='right')
    # print(merge6A)
    merge6A.replace('', np.nan, inplace=True)
    merge6A.dropna(subset=['CounterA'], inplace=True)
    merge6A.dropna(subset=['CounterB'], inplace=True)
    # print(merge6A)
    # print(merge6A.head())
    #                      name   DC    AC Contract  CounterB  CounterA
    # 0        Vodafone B02.104  703  7500   143241  30101318  30102704
    # 1          R1954 Vodafone   20  2000   141155  30101762  30101741
    # 2  R0428 Kvant-Telecom AO    1  3500   142954  30102685  30100632
    # 3        R0408 Data Group  125  1000   142989  30103976  30100546
    # 4     R0401 Truphone Ltd.    6   600   142683  30101027  30100372

    merge7A = pd.merge(merge6A, mysqlDF, left_on='CounterA',
                       right_on='Counter', how='right')
    # print(merge7A.head())
    # print(mysqlDF.head())

    mysqlDF.rename({"Counter": "CounterB", "name": "nameB", "Rack": "RackB", "Month": "MonthB", "Usage": "UsageB"},axis='columns',inplace=True)

    # print(mysqlDF.head())
    # print(merge6A.head())
    merge7B = pd.merge(merge6A, mysqlDF, left_on='CounterB',
                       right_on='CounterB', how='right')
    # print(merge7B.head())
    # print(merge7B)
    # print(merge7A.head())
#                name_x       DC    AC Contract  CounterB  CounterA   Counter                  name_y   Rack   Month   Usage
# 0        Vodafone B02.104  703  7500   143241  30101318  30102704  30102704  RB Vodafone Enterprise     RB  201911     0.0
# 1  R0428 Kvant-Telecom AO    1  3500   142954  30102685  30100632  30100632  R0428 Kvant-Telecom AO  R0428  201911  1259.0
# 2        R0408 Data Group  125  1000   142989  30103976  30100546  30100546         R0408 DataGroup  R0408  201911   138.0
# 3     R0401 Truphone Ltd.    6   600   142683  30101027  30100372  30100372      R0401 Truphone Ltd  R0401  201911     6.0
# 4              R2733 Retn   63  7000   140528  30100845  30100778  30100778         R2733 Retn GmbH  R2733  201911   224.0
    # print(merge7B.head())
#              name_x         DC    AC Contract  CounterB  CounterA   Counter                  name_y   Rack   Month   Usage
# 0        Vodafone B02.104  703  7500   143241  30101318  30102704  30101318  RB Vodafone Enterprise     RB  201911     0.0
# 1  R0428 Kvant-Telecom AO    1  3500   142954  30102685  30100632  30102685  R0428 Kvant-Telecom AO  R0428  201911  1228.0
# 2        R0408 Data Group  125  1000   142989  30103976  30100546  30103976         R0408 DataGroup  R0408  201911   106.0
# 3     R0401 Truphone Ltd.    6   600   142683  30101027  30100372  30101027      R0401 Truphone Ltd  R0401  201911     0.0
# 4           R2722 JSC TTK   32  2000   142993  30111740  30100071  30111740  R2722 JSC TransTelecom  R2722  201911     0.0

    # print(merge7A.head())
    merge7A = merge7A.drop('Counter', 1)
    merge7A = merge7A.drop('name_y', 1)
    merge7A = merge7A.rename(columns={'Usage': 'Usage_A'})
    # print(merge7A.loc[merge7A['name_x'] == 'R2004 Cogent'])
    # print()
    # merge7A['Usage_A'] = merge7A['Usage_A'].infer_objects()
    # print(merge7A.head())
    # print(merge7A.info())
    #                    name_x   DC    AC Contract  CounterB  CounterA   Rack   Month  Usage_A
    # 0        Vodafone B02.104  703  7500   143241  30101318  30102704     RB  201911      0.0
    # 1  R0428 Kvant-Telecom AO    1  3500   142954  30102685  30100632  R0428  201911   1259.0
    # 2        R0408 Data Group  125  1000   142989  30103976  30100546  R0408  201911    138.0
    # 3     R0401 Truphone Ltd.    6   600   142683  30101027  30100372  R0401  201911      6.0
    # 4              R2733 Retn   63  7000   140528  30100845  30100778  R2733  201911    224.0

    # print(merge7B.head())
    # merge7B = merge7B.drop('Counter', 1)
    merge7B = merge7B.drop('nameB', 1)
    merge7B = merge7B.rename(columns={'UsageB': 'Usage_B'})
    # merge7B['Usage_B'] = merge7B['Usage_B'].infer_objects()
    # filter = merge7B["name"] == "R2004 Cogent"
    # print(merge7B.where(filter, inplace = False).head())
    # print(merge7B.loc[merge7B['name'] == 'R2004 Cogent'])
    # print()
    # print(merge7B.info())
    #                    name_x   DC    AC Contract  CounterB  CounterA   Rack   Month  Usage_B
    # 0        Vodafone B02.104  703  7500   143241  30101318  30102704     RB  201911      0.0
    # 1  R0428 Kvant-Telecom AO    1  3500   142954  30102685  30100632  R0428  201911   1228.0
    # 2        R0408 Data Group  125  1000   142989  30103976  30100546  R0408  201911    106.0
    # 3     R0401 Truphone Ltd.    6   600   142683  30101027  30100372  R0401  201911      0.0
    # 4           R2722 JSC TTK   32  2000   142993  30111740  30100071  R2722  201911      0.0

    merge7C = pd.merge(merge7B, merge7A[['CounterA', 'Usage_A']],
                       left_index=False, right_index=True, on='CounterA', how='right')

    # print(merge7C.loc[merge7C['name'] == 'R2004 Cogent'])
    # merge7B = merge7B.drop_duplicates().sort_values(by='Contract')
    # print(merge7C.head())
    #                    name_x   DC    AC Contract  CounterB  CounterA   Rack   Month  Usage_B  Usage_A
    # 0        Vodafone B02.104  703  7500   143241  30101318  30102704     RB  201911      0.0      0.0
    # 1  R0428 Kvant-Telecom AO    1  3500   142954  30102685  30100632  R0428  201911   1228.0   1259.0 (CHECKED - CORRECT)
    # 2        R0408 Data Group  125  1000   142989  30103976  30100546  R0408  201911    106.0    138.0
    # 3     R0401 Truphone Ltd.    6   600   142683  30101027  30100372  R0401  201911      0.0      6.0
    # 4           R2722 JSC TTK   32  2000   142993  30111740  30100071  R2722  201911      0.0    224.0

    # print(merge7B.info())
    # print(merge7B)
    # merge7B = merge7A.drop_duplicates().sort_values(by='name')
    # print(merge7A.head())

    return [merge7C, mysqlm1DF, mysqlm2DF, date]


def sendMail(date, merge7):
    year = date[0:4]
    month = date[-2:]
    arrowDate = arrow.get(int(year), int(month), 1)
    formattedDate = arrowDate.format('MMMM YYYY')
    formattedMonth = arrowDate.format('MMMM')
    print('to: ndomino@newtelco.de')
    # print('cc: izhuravel@newtelco.de')
    # print('cc: sburtsev@newtelco.de')
    # print('cc: power@newtelco.de')
    # print('cc: billing@newtelco.de')
    # print('cc: order@newtelco.de')
    # print('cc: sales@newtelco.de')
    print('From: device@newtelco.de')
    print('MIME-Version: 1.0')
    print('Content-Type: multipart/mixed; boundary=multipart-boundary')
    print('Subject: [POWER USAGE] Monthly Power Comparison (' + formattedDate + ')')
    print('--multipart-boundary')
    print('Content-Type: text/html; charset=utf-8')
    print('')
    print('<html>')
    print('<head>')
    print('<style>')
    print(
        'resultTable { border-spacing: 20px; font-size: 1.2rem; text-align: center; }')
    print('body { font-size: 1.3rem; }')
    print('td, th { padding: 5px !important; }')
    print('</style>')
    print('</head>')
    print('<body>')
    print('<p>')
    print('Dear Colleagues,')
    print('<br/><br/>')
    print('Below is the power usage comparison for <b>' + formattedDate + '</b>')
    print('<br/><br/>')
    print('Please see the Excel Attachment for more in-depth data')
    print('<br/><br/>')
    print('This is a beta version, if you find any errors - please report them to ndomino@newtelco.de')
    print('</p>')
    print('')
    print('<hr style="max-width: 500px; margin-right: 70%;" />')
    print('')

    # Days in the month for calculating later on
    monthsArray = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    #               J   F   M   A   M   J   J   A   S   O   N   D
    # Leap Year
    #
    # if (YEAR % 4) === 0:
    #   is leap-year
    #

    rackAC0 = merge7.filter(['Rack', 'Contract', 'AC'],
                            axis=1).drop_duplicates()
    rackAC0['AC'] = pd.to_numeric(rackAC0['AC'])
    rackAC0 = rackAC0.sort_values(by='Contract')
    rackAC0 = rackAC0.groupby(['Contract'])['AC'].sum()

    merge81 = pd.merge(merge7, rackAC0, left_on='Contract',
                       right_on='Contract', how='left')
    merge81 = merge81.sort_values(by='Contract')
    merge81 = merge81.drop(['AC_x'], axis=1)
    # merge81 = merge81.drop(['DC'], axis=1)
    merge81.rename(columns={'name': 'Name','RackB': 'Rack', 'MonthB': 'Month', 'Usage_B': 'Usage B-Feed (kWh)', 'Usage_A': 'Usage A-Feed (kWh)',
                            'Total_Usage': 'Total Usage (kWh)', 'AC_y': 'Allowed AC Usage (W)'}, inplace=True)
    merge81['Month'] = formattedMonth
    # print(merge81.head())
    # print(merge81.loc[merge81['Contract'] == '143106'])
    merge82 = merge81.groupby(['Contract'])

    for name, group in merge82:
        diffSum = pd.to_numeric(group['Total Usage (kWh)']).sum()
        # print(name)
        # if name == '143106':
            # print(diffSum)
        monthValue = int(date[-2:])
        monthValue -= 1
        monthHrs = monthsArray[int(monthValue)]
        monthHrs = int(monthHrs * 24)
        # diffSum = (diffSum * monthHrs)
        diffSum = truncate(diffSum, 2)
        groupAC = group['Allowed AC Usage (W)'].max()
        if str(group['Allowed AC Usage (W)'].max()) != 'nan' and str(group['Allowed AC Usage (W)'].max()) != '0.0':
            avgAC = ((group['Allowed AC Usage (W)'].sum() / 1000) * monthHrs)
            diffAC = int(avgAC) - int(diffSum)
        if diffAC < 0:
            print('<p>')
            print('<h4>Contract: <font style="weight:700">' +
                  name + '</font></h4>')
            print(group.to_html(classes='resultTable', index=False, border=1,
                                bold_rows=True, na_rep='-', justify='center'))
            print('')
            # diffSum = (diffSum)
            # avgAC = (avgAC)
            diffSum = float("{0:.2f}".format(diffSum))
            print('<br/>')
            print('Monthly Usage: ' + str(diffSum) + ' kWh')
            print('<br/>')
            print('Allowed Usage: ' + str(avgAC) + ' kWh')
            print('<br/>')
            print('<br/>')
            diffAC = diffAC * -1
            diffAC = float("{0:.2f}".format(diffAC))
            diffKw = diffAC / monthHrs
            diffKw = float("{0:.2f}".format(diffKw))
            if diffAC > 100:
                print('<font>Over Usage (Überverbrauch): ' +
                      str(diffAC) + ' kWh</font><br>')
                print('<font style="color:red;font-weight:700">Over Usage (Überverbrauch): ' +
                      str(diffKw) + ' kW</font><br>')
            else:
                print('<font>Over Usage (Überverbrauch): ' +
                      str(diffAC) + ' kWh</font><br>')
                print('<font style="color:red;font-weight:700">Over Usage (Überverbrauch): ' +
                      str(diffKw) + ' kW</font><br>')
            # if str(group['DC'].max()) != 'nan':
            #     avgDC = group['DC'].max()
            #     print('<br/><br/>')
            #     print('Allowed Usage DC: ' + str(avgDC))
            #     print('<br/>')
            #     print('Over Usage DC (Überverbrauch): ' + ' Watt ')
            # print('---------------------' + '<br>')
            print('</p>')
            print('<hr style="max-width: 500px; margin-right: 70%;" />')

    print('</body>')
    print('</html>')
    print('--multipart-boundary')


def createWorksheet(primaryData, mysqlm1DF, mysqlm2DF, date):
    wb = Workbook()
    ws = wb.active

    # Begin merging separate query / column data into one
    primaryData['Total_Usage'] = primaryData['Usage_A'] + \
        primaryData['Usage_B']
    merge9 = pd.merge(primaryData, mysqlm1DF,
                      left_on='CounterB', right_on='Counter', how='left')
    merge9A = pd.merge(merge9, mysqlm1DF, left_on='CounterA',
                       right_on='Counter', how='left')
    merge9A[['Usage M-1_x', 'Usage M-1_y']
            ] = merge9A[['Usage M-1_x', 'Usage M-1_y']].apply(pd.to_numeric)
    merge9A = merge9A.drop(['Month-1_y'], axis=1)
    merge9A = merge9A.drop(['Counter_x'], axis=1)
    merge9A = merge9A.drop(['Counter_y'], axis=1)
    merge9A['Total_Usage_M1'] = merge9A['Usage M-1_x'] + merge9A['Usage M-1_y']

    merge10 = pd.merge(merge9A, mysqlm2DF, left_on='CounterB',
                       right_on='Counter', how='left')
    merge10A = pd.merge(merge10, mysqlm2DF, left_on='CounterA',
                        right_on='Counter', how='left')
    merge10A[['Usage M-2_x', 'Usage M-2_y']
             ] = merge10A[['Usage M-2_x', 'Usage M-2_y']].apply(pd.to_numeric)
    merge10A = merge10A.drop(['Month-2_y'], axis=1)
    merge10A = merge10A.drop(['Counter_x'], axis=1)
    merge10A = merge10A.drop(['Counter_y'], axis=1)
    merge10A['Total_Usage_M2'] = merge10A['Usage M-2_x'] + \
        merge10A['Usage M-2_y']

    merge10A = merge10A.drop_duplicates().sort_values(by=['Contract', 'RackB'])
    merge10A['contractDiff'] = merge10A['Contract'] == merge10A['Contract'].shift(
        1).fillna(merge10A['Contract'])
    merge10A = merge10A[pd.notnull(merge10A["Contract"])]

    # Calculate hours in month for next calculation (w -> wh)
    monthsArray = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    monthValue = int(date[-2:])
    monthValue -= 1
    monthHrs = monthsArray[int(monthValue)]
    monthHrs = monthHrs * 24

    merge10A.insert(11, 'O0', '')
    merge10A.insert(16, 'O1', '')
    merge10A.insert(21, 'O2', '')

    merge10A = merge10A.rename(columns={'Total_Usage': 'Total_Usage_M0'})

    # Calculate Usage in Wh from W
    for row in merge10A.itertuples():
        merge10A.at[row.Index, 'O0'] = float(
            row.Total_Usage_M0)
        merge10A.at[row.Index, 'O1'] = float(
            row.Total_Usage_M1)
        merge10A.at[row.Index, 'O2'] = float(
            row.Total_Usage_M2)

    merge10A.reset_index(inplace=True)

    merge10A['O0'] = pd.to_numeric(merge10A['O0'])
    merge10A['O1'] = pd.to_numeric(merge10A['O1'])
    merge10A['O2'] = pd.to_numeric(merge10A['O2'])

    # Sum Month 0 Usage in kwH per Contract
    merge10A.insert(13, 'Contract_CumSum_M0', '')
    merge10A['Contract_CumSum_M0'] = merge10A.groupby(['Contract'])[
        'O0'].cumsum()

    contractSum0 = merge10A.groupby(['Contract'])[
        'Contract_CumSum_M0'].agg('max')

    merge11 = pd.merge(merge10A, contractSum0, how='left',
                       left_on='Contract', right_on='Contract')

    # Sum Month 1 Usage in kwH per Contract
    merge11.insert(20, 'Contract_CumSum_M1', '')

    merge11['Contract_CumSum_M1'] = merge11.groupby(['Contract'])[
        'O1'].cumsum()

    contractSum1 = merge11.groupby(['Contract'])[
        'Contract_CumSum_M1'].agg('max')

    merge12 = pd.merge(merge11, contractSum1, how='left',
                       left_on='Contract', right_on='Contract')

    # Sum Month 2 Usage in kwH per Contract
    merge12.insert(27, 'Contract_CumSum_M2', '')
    merge12['Contract_CumSum_M2'] = merge12.groupby(['Contract'])[
        'O2'].cumsum()

    contractSum2 = merge12.groupby(['Contract'])[
        'Contract_CumSum_M2'].agg('max')

    merge13 = pd.merge(merge12, contractSum2, how='left',
                       left_on='Contract', right_on='Contract')

    # Calculate total allowance per Contract (sum of rack allowance)
    # print(merge13[['AC']])
    # pd.set_option('display.max_rows', None)
    # print(merge13)
    merge13[['AC']] = merge13[['AC']].apply(pd.to_numeric)

    contractSum2 = merge13.groupby(['Contract'])[
        'AC'].agg('sum')

    merge13 = pd.merge(merge13, contractSum2, how='left',
                       left_on='Contract', right_on='Contract')

    # Calculate Contract Overages
    merge13['Overage_M0'] = 0
    merge13['Overage_M1'] = 0
    merge13['Overage_M2'] = 0

    for row in merge13.itertuples():
        merge13.at[row.Index, 'Overage_M0'] = row.AC_y - float(
            row.Contract_CumSum_M0_y)
        merge13.at[row.Index, 'Overage_M1'] = row.AC_y - float(
            row.Contract_CumSum_M1_y)
        merge13.at[row.Index, 'Overage_M2'] = row.AC_y - float(
            row.Contract_CumSum_M2_y)

    # Organise and rename columns
    merge13 = merge13.rename(columns={'O0': 'Usage_wH_0'})
    merge13 = merge13.rename(columns={'O1': 'Usage_wH_1'})
    merge13 = merge13.rename(columns={'O2': 'Usage_wH_2'})

    merge13 = merge13.rename(
        columns={'Contract_CumSum_M0_x': 'Contract_Usage_wH_0'})
    merge13 = merge13.rename(
        columns={'Contract_CumSum_M1_x': 'Contract_Usage_wH_1'})
    merge13 = merge13.rename(
        columns={'Contract_CumSum_M2_x': 'Contract_Usage_wH_2'})

    merge13 = merge13.rename(
        columns={'AC_x': 'RackAllowance_AC'})
    merge13 = merge13.rename(
        columns={'AC_y': 'ContractAllowance_AC'})

    # Begin Excel Worksheet Manipulation
    excelRows = dataframe_to_rows(merge13)

    for r_idx, row in enumerate(excelRows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws.insert_cols(7)
    ws.insert_cols(17, amount=2)
    ws.insert_cols(24, amount=2)
    ws.insert_cols(32, amount=2)
    # ws.insert_cols(24)
    # ws.insert_cols(25)

    i = 1
    rowCount = ws.max_row
    while i <= rowCount:
        move_cell(ws['AM' + str(i)], 'G' + str(i), ws)
        move_cell(ws['AI' + str(i)], 'Q' + str(i), ws)
        move_cell(ws['AN' + str(i)], 'R' + str(i), ws)
        move_cell(ws['AK' + str(i)], 'X' + str(i), ws)
        move_cell(ws['AO' + str(i)], 'Y' + str(i), ws)
        move_cell(ws['AL' + str(i)], 'AF' + str(i), ws)
        move_cell(ws['AP' + str(i)], 'AG' + str(i), ws)
        i += 1

    ws.column_dimensions['AJ'].hidden = True
    ws.column_dimensions['AI'].hidden = True
    ws.column_dimensions['AH'].hidden = True
    ws.column_dimensions['AA'].hidden = True
    ws.column_dimensions['P'].hidden = True
    ws.column_dimensions['B'].hidden = True
    ws.column_dimensions['D'].hidden = True

    # Excel Header Names
    ws['C1'] = 'Rack Name'
    ws['E1'] = 'Rack Allowed (W)'
    ws['F1'] = 'Contract #'
    ws['G1'] = 'Contract Allowance (W)'
    ws['H1'] = 'Counter # (B-Feed)'
    ws['I1'] = 'Counter # (A-Feed)'
    ws['J1'] = 'Rack #'
    ws['K1'] = 'Month 1'
    ws['L1'] = 'Usage B-Feed (kWh)'
    ws['M1'] = 'Usage A-Feed (kWh)'
    ws['N1'] = 'Total (kWh) M1'
    ws['O1'] = 'XXX Total (kWh) M1'
    ws['Q1'] = 'Contract (Wh) M1'
    ws['R1'] = 'Overage M1'
    ws['S1'] = 'Month 2'
    ws['T1'] = 'Usage A-Feed (W)'
    ws['U1'] = 'Usage B-Feed (W)'
    ws['V1'] = 'Total (W) M2'
    ws['W1'] = 'Total (Wh) M2'
    ws['X1'] = 'Contact (Wh) M2'
    ws['Y1'] = 'Overage M2'
    ws['Z1'] = 'Month 3'
    ws['AB1'] = 'Usage A-Feed (W)'
    ws['AC1'] = 'Usage B-Feed (W)'
    ws['AD1'] = 'Total (W) M3'
    ws['AE1'] = 'Total (Wh) M3'
    ws['AF1'] = 'Contract (Wh) M3'
    ws['AG1'] = 'Overage M3'

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max(
                    (dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 3

    for row in range(3, ws.max_row + 1):
        for col in range(2, ws.max_column):
            ws.cell(column=col, row=row).font = Font(
                color="000000", name="Calibri", size="12")

    for cell in ws['L']:
        cell.style = 'Comma'
    for cell in ws['M']:
        cell.style = 'Comma'
    for cell in ws['N']:
        cell.style = 'Comma'
    for cell in ws['O']:
        cell.style = 'Comma'
    for cell in ws['Q']:
        cell.style = 'Comma'
    for cell in ws['R']:
        cell.style = 'Comma'
    for cell in ws['T']:
        cell.style = 'Comma'
    for cell in ws['U']:
        cell.style = 'Comma'
    for cell in ws['V']:
        cell.style = 'Comma'
    for cell in ws['W']:
        cell.style = 'Comma'
    for cell in ws['X']:
        cell.style = 'Comma'
    for cell in ws['Y']:
        cell.style = 'Comma'
    for cell in ws['AB']:
        cell.style = 'Comma'
    for cell in ws['AC']:
        cell.style = 'Comma'
    for cell in ws['AD']:
        cell.style = 'Comma'
    for cell in ws['AE']:
        cell.style = 'Comma'
    for cell in ws['AF']:
        cell.style = 'Comma'
    for cell in ws['AG']:
        cell.style = 'Comma'

    i = 3
    rowCount = ws.max_row
    while i <= rowCount:
        if float(ws['Y' + str(i)].value) < 0.0:
            ws['Y' + str(i)].style = 'Bad'
        if int(ws['R' + str(i)].value) < 0:
            ws['R' + str(i)].style = 'Bad'
        if float(ws['AG' + str(i)].value) < 0.0:
            ws['AG' + str(i)].style = 'Bad'
        i += 1

    key_column = 6
    merge_columns = [6, 7, 17, 18, 24, 25, 32, 33]
    start_row = 3
    max_row = ws.max_row
    key = None

    # Iterate all rows in `key_colum`
    for row, row_cells in enumerate(ws.iter_rows(min_col=key_column, min_row=start_row, max_col=key_column, max_row=max_row + 1), start_row):
        if key != row_cells[0].value or row == max_row:
            # moved line below this if
            # if row == max_row: row += 1
            if not key is None:
                for merge_column in merge_columns:
                    ws.merge_cells(start_row=start_row, start_column=merge_column,
                                   end_row=row - 1, end_column=merge_column)

                    ws.cell(row=start_row, column=merge_column).\
                        alignment = Alignment(
                            horizontal='center', vertical='center')

                start_row = row

            key = row_cells[0].value
        # moved below line here as it was merging last two rows content even if the values differ.
        if row == max_row:
            row += 1

    rh1 = ws.row_dimensions[1]
    rh1.height = 25

    for cell in ws['A'] + ws[1]:
        cell.style = 'Headline 2'
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Excel Formatting
    bottomBorder = NamedStyle(name="bottomBorder")
    bottomBorder.alignment = Alignment(horizontal="center")
    bottomBorder.alignment = Alignment(vertical="center")
    bottomBorder.alignment = Alignment(wrapText=1)
    bottomBorder.font = Font(
        color="ffffff", name="Calibri", size="12", bold=True)
    bottomBorder.fill = PatternFill("solid", fgColor="67B246")
    bd = Side(style='thin', color="ffffff")
    bottomBorder.border = Border(bottom=bd, right=bd, left=bd)

    for cell in ws['A'] + ws[1]:
        cell.style = bottomBorder
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # for cell in ws
    ws.sheet_view.zoomScale = 75

    # Excel Column Filtering
    FullRange = "B1:" + \
        get_column_letter(ws.max_column) + str(ws.max_row)
    ws.auto_filter.ref = FullRange

    # Freeze Row 1 / Column A
    c = ws['B2']
    ws.freeze_panes = c

    filenameDate = datetime.datetime.now().strftime("%d%m%Y")
    wb.save("/opt/newtelco/powerCompare/output/excel/powerCompare_" +
            date + "_" + filenameDate + ".xlsx")


def main(argv):
    selectedDate = ''
    sendMailB = 1

    try:
        opts, args = getopt.getopt(argv, "hd:m", ["help", "date=", "sendmail"])
    except getopt.GetoptError:
        print('Usage: power.py -d <date i.e. 201808>')
        sys.exit(2)
    for opt, arg in opts:
        if opt in ('-h', '--help'):
            print('Usage: power.py -d <date i.e. 201808>')
            sys.exit()
        elif opt in ("-d", "--date"):
            selectedDate = arg
        elif opt in ("-m", "--sendmail"):
            sendMailB = 0

    primaryData, mysqlm1DF, mysqlm2DF, date = compare(selectedDate)

    if sendMailB == 0:
        createWorksheet(primaryData, mysqlm1DF, mysqlm2DF, date)
        sendMail(date, primaryData)
    else:
        createWorksheet(primaryData, mysqlm1DF, mysqlm2DF, date)


if __name__ == "__main__":
    main(sys.argv[1:])
